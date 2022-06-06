#!/usr/bin/env python3
# encoding: utf-8
'''
asap.prepareJSONInput -- Create a JSON input file for ASAP from a multifasta or Excel spreadsheet

asap.prepareJSONInput

@author:     Darrin Lemmer

@copyright:  2015,2019 TGen North. All rights reserved.

@license:    ACADEMIC AND RESEARCH LICENSE -- see ../LICENSE

@contact:    dlemmer@tgen.org
'''

import sys
import os
import re
import argparse
import logging
import skbio.io
from skbio import DNA
from openpyxl import load_workbook

from asap import dispatcher
from asap import assayInfo
from asap import __version__

__all__ = []
__updated__ = '2019-01-15'
__date__ = '2015-08-03'

DEBUG = 1
TESTRUN = 0
PROFILE = 0

PRESENCE_ABSENCE = 10
GENE_VARIANT = 20

def _process_fasta(fasta, fasta_type, message=None):
    return_list = []
    for seq in skbio.io.registry.read(fasta, format='fasta', constructor=DNA):
        significance = assayInfo.Significance(seq.metadata['description']) if seq.metadata['description'] else message
        amplicon = assayInfo.Amplicon(sequence=_clean_seq(str(seq)), significance=significance)
        if fasta_type == GENE_VARIANT:
            amplicon.variant_name = seq.metadata['id']
            return_list.append(amplicon)
        else: #PRESENCE_ABSENCE
            target = assayInfo.Target(function='species ID', amplicon=amplicon) #'species ID' is a guess here, TODO: make this smarter
            assay = assayInfo.Assay(name=seq.metadata['id'], assay_type='presence/absence', target=target)
            return_list.append(assay)
    return return_list

def _process_fasta_single(fasta):
    for seq in skbio.io.registry.read(fasta, format='fasta', constructor=DNA):
        amplicon = assayInfo.Amplicon(sequence=_clean_seq(str(seq)))
    return amplicon

def _clean_seq(sequence):
    return_seq = sequence.upper()
    return_seq = re.sub('[-|_]', '', return_seq)
    return return_seq

def _clean_str(string):
    if string:
        return re.sub(' ', '_', string)
    else:
        return None

def _strip(string):
    if string or string is 0:
        if type(string) is str:
            return string.strip()
        else:
            return str(string)
    else:
        return None

def _isNT(sequence, positions):
    size = 0
    for token in positions.split(','):
        m = re.search(r"(\d*)-(\d*)", token)
        if m:
            size += int(m.group(2)) - int(m.group(1)) + 1
        else:
            size += 1
    if size == len(sequence):
        return True
    else:
        return False

class CLIError(Exception):
    '''Generic exception to raise and log different fatal errors.'''
    def __init__(self, msg):
        super(CLIError).__init__(type(self))
        self.msg = "E: %s" % msg
    def __str__(self):
        return self.msg
    def __unicode__(self):
        return self.msg

def main(argv=None): # IGNORE:C0111
    '''Command line options.'''

    if argv is None:
        argv = sys.argv
    else:
        if not isinstance(argv, argparse.Namespace):
            sys.argv.extend(argv)
            pass

    program_name = os.path.basename(sys.argv[0])
    program_version = "v%s" % __version__
    program_build_date = str(__updated__)
    program_version_message = '%%(prog)s %s (%s)' % (program_version, program_build_date)
    if __name__ == '__main__':
        program_shortdesc = __import__('__main__').__doc__.split("\n")[1]
    else:
        program_shortdesc = __doc__.split("\n")[1]
    program_license = '''%s

  Created by TGen North on %s.
  Copyright 2015 TGen North. All rights reserved.

  Available for academic and research use only under a license
  from The Translational Genomics Research Institute (TGen)
  that is free for non-commercial use.

  Distributed on an "AS IS" basis without warranties
  or conditions of any kind, either express or implied.

USAGE
''' % (program_shortdesc, str(__date__))

    try:
        # Setup argument parser
        # parser = argparse.ArgumentParser(description=program_license, formatter_class=argparse.RawDescriptionHelpFormatter)
        # required_group = parser.add_argument_group("required arguments")
        # exclusive_group = required_group.add_mutually_exclusive_group(required=True)
        # exclusive_group.add_argument("-f", "--fasta", metavar="FILE", help="fasta file containing amplicon sequences.")
        # exclusive_group.add_argument("-x", "--excel", metavar="FILE", help="Excel file of assay data.")
        # required_group.add_argument("-o", "--out", metavar="FILE", required=True, help="output JSON file to write. [REQUIRED]")
        # parser.add_argument("-w", "--worksheet", help="Excel worksheet to use, the first one in the file will be used if not specified")
        # parser.add_argument('-v', '--version', action='version', version=program_version_message)

        # Process arguments
        if isinstance(argv, argparse.Namespace):
            args = argv
            pass
        else:
            args = asapParser.parser.parse_args(argv)

        fasta_file = args.fasta
        excel_file = args.excel
        out_file = args.out
        worksheet = args.worksheet

        assay_list = []

        if fasta_file:
            assay_list = _process_fasta(fasta_file, PRESENCE_ABSENCE)
        else:
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active
            if worksheet:
                ws = wb[worksheet]
            amplicon = None
            target = None
            assay = None
            for row in ws.iter_rows(row_offset=2):

                if _strip(row[0].value): #Create a new assay
                    if assay:
                        assay_list.append(assay)
                        target = None
                        amplicon = None
                    assay = assayInfo.Assay(name=_clean_str(_strip(row[0].value)), assay_type=_strip(row[1].value))

                if _strip(row[18].value):
                    significance = assayInfo.Significance(message=_strip(row[17].value), resistance=_strip(row[18].value))
                else:
                    significance = assayInfo.Significance(message=_strip(row[17].value))
                element = None
                if _strip(row[14].value): #Significance gets attached to ROI
                    sequence = _strip(row[15].value)
                    positions = _strip(row[14].value)
                    if _isNT(sequence, positions):
                        element = assayInfo.RegionOfInterest(position_range=positions, nt_sequence=sequence, mutations=_strip(row[16].value), name=_strip(row[13].value), significance=significance)
                    else:
                        element = assayInfo.RegionOfInterest(position_range=positions, aa_sequence=sequence, mutations=_strip(row[16].value), name=_strip(row[13].value), significance=significance)
                elif _strip(row[10].value): #Significance gets attached to SNP
                    element = assayInfo.SNP(position=_strip(row[10].value), reference=_strip(row[11].value), variant=_strip(row[12].value), name=_strip(row[9].value), significance=significance)
                if _strip(row[8].value): #Process amplicon
                    if os.path.isfile(_strip(row[8].value)):
                        if assay.assay_type == "gene variant":
                            amplicon = _process_fasta(_strip(row[8].value), GENE_VARIANT, significance)
                        else:
                            amplicon = _process_fasta_single(_strip(row[8].value))
                            if element:
                                amplicon.add_SNP(element) if isinstance(element, assayInfo.SNP) else amplicon.add_ROI(element)
                            else:
                                amplicon.significance = significance
                    else:
                        amplicon = assayInfo.Amplicon(sequence=_clean_seq(_strip(row[8].value)), variant_name=_clean_str(_strip(row[7].value)))
                        if element:
                            amplicon.add_SNP(element) if isinstance(element, assayInfo.SNP) else amplicon.add_ROI(element)
                        else:
                            amplicon.significance = significance
                elif amplicon and element: #We already have an amplicon, but we need to add another SNP or ROI to it
                   amplicon.add_SNP(element) if isinstance(element, assayInfo.SNP) else amplicon.add_ROI(element)

                if target and _strip(row[8].value):
                    target.add_amplicon(amplicon)
                    amplicon = None
                elif target:
                    target.amplicon = amplicon
                else:
                    target = assayInfo.Target(function=_strip(row[2].value), gene_name=_strip(row[3].value), start_position=_strip(row[4].value), end_position=_strip(row[5].value), reverse_comp=_strip(row[6].value), amplicon=amplicon)
                    assay.target = target

            assay_list.append(assay) #get the last one

        assay_data = {"Assay":assay_list}
        assayInfo.writeJSON(assay_data, out_file)

        return 0
    except KeyboardInterrupt:
        ### handle keyboard interrupt ###
        return 0
    except Exception as e:
        if DEBUG or TESTRUN:
            raise(e)
        indent = len(program_name) * " "
        sys.stderr.write(program_name + ": " + repr(e) + "\n")
        sys.stderr.write(indent + "  for help use --help")
        return 2

if __name__ == "__main__":
    if DEBUG:
        pass
    if TESTRUN:
        import doctest
        doctest.testmod()
    if PROFILE:
        import cProfile
        import pstats
        profile_filename = 'asap.outputCombiner_profile.txt'
        cProfile.run('main()', profile_filename)
        statsfile = open("profile_stats.txt", "wb")
        p = pstats.Stats(profile_filename, stream=statsfile)
        stats = p.strip_dirs().sort_stats('cumulative')
        stats.print_stats()
        statsfile.close()
        sys.exit(0)
    sys.exit(main())
